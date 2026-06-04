"""Regenerate W20 Excel with correct 17 columns."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

INPUT = r"D:\Baek04w\待做\每周文献调研(1)\每周文献调研\中间结果\step03-classified-2026-W20.md"
OUTPUT = r"D:\Baek04w\待做\每周文献调研(1)\每周文献调研\文献调研\文献调研记录_2026-W20.xlsx"

# Read file
with open(INPUT, "r", encoding="utf-8") as f:
    content = f.read()

# Find main table (starts after line with "| # | 标题 | ...)
table_start = content.find("| # | 标题 |")
table_end = content.find("\n---\n", table_start)
table_lines = content[table_start:table_end].strip().split("\n")

# Header row
headers = [h.strip() for h in table_lines[0].split("|")[1:-1]]
print("Headers:", headers)

# Data rows (skip header, separator)
rows = []
for line in table_lines[2:]:
    if line.startswith("|---") or not line.strip():
        continue
    cells = [c.strip() for c in line.split("|")[1:-1]]
    rows.append(cells)

print(f"Rows: {len(rows)}")

# Column mapping: step03 header index -> output column index
# step03 headers: #, 标题, 作者, 期刊, 发表日期, DOI, 发表状态, 摘要(完整), 论文关键词, 核心发现, 检索来源, 关键词编号, 检索组, 研究方向, 优先级, JCR分区, 影响因子, 备注
# output: 序号, 标题, 作者, 期刊, 发表日期, 研究方向, 核心发现, 摘要, 关键词, DOI, 发表状态, JCR分区, 影响因子, 检索来源, 关键词编号, 优先级, 备注
STEP03_IDX = {
    "序号": 0,
    "标题": 1,
    "作者": 2,
    "期刊": 3,
    "发表日期": 4,
    "DOI": 5,
    "发表状态": 6,
    "摘要": 7,
    "关键词": 8,
    "核心发现": 9,
    "检索来源": 10,
    "关键词编号": 11,
    "检索组": 12,
    "研究方向": 13,
    "优先级": 14,
    "JCR分区": 15,
    "影响因子": 16,
    "备注": 17,
}

OUTPUT_COLS = ["序号", "标题", "作者", "期刊", "发表日期", "研究方向", "核心发现",
               "摘要", "关键词", "DOI", "发表状态", "JCR分区", "影响因子",
               "检索来源", "关键词编号", "优先级", "备注"]

def get_val(row, key):
    idx = STEP03_IDX[key]
    if idx < len(row):
        v = row[idx]
        # Truncate abstract to 500 chars
        if key == "摘要" and len(v) > 500:
            v = v[:500] + "..."
        return v
    return ""

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "文献调研-W20"

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill("solid", fgColor="D9E1F2")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Write headers
for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

# Write data rows
for row_idx, row in enumerate(rows, 2):
    for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
        val = get_val(row, col_name)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if col_name in ["摘要", "核心发现", "标题"]:
            cell.alignment = left_wrap
        else:
            cell.alignment = center

# Set column widths
col_widths = {
    "序号": 6, "标题": 35, "作者": 28, "期刊": 18, "发表日期": 12,
    "研究方向": 22, "核心发现": 40, "摘要": 60, "关键词": 30,
    "DOI": 25, "发表状态": 14, "JCR分区": 10, "影响因子": 10,
    "检索来源": 22, "关键词编号": 12, "优先级": 10, "备注": 18
}
for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
    ws.column_dimensions[chr(64 + col_idx)].width = col_widths.get(col_name, 15)

# Freeze header
ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Rows written: {len(rows)}")

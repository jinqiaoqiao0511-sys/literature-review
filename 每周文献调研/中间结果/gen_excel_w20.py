#!/usr/bin/env python3
"""Generate Excel file from step03-classified markdown for weekly literature review."""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Config
ISO_WEEK = "2026-W20"
INPUT_FILE = r"D:\Baek04w\待做\每周文献调研(1)\每周文献调研\中间结果\step03-classified-2026-W20.md"
OUTPUT_FILE = rf"D:\Baek04w\待做\每周文献调研(1)\每周文献调研\文献调研\文献调研记录_{ISO_WEEK}.xlsx"
SURVEY_DATE = "2026-05-20"

# Read input
with open(INPUT_FILE, "r", encoding="utf-8") as f:
    content = f.read()

# Parse markdown table rows
# Find the table section
table_match = re.search(r'\| # \|.*?\n\|---\|.*?\n([\s\S]*?)(?=\n---|\n## |\Z)', content)
if not table_match:
    print("ERROR: Could not find table in input file")
    exit(1)

table_body = table_match.group(1).strip()
rows = []
for line in table_body.split("\n"):
    line = line.strip()
    if not line or not line.startswith("|"):
        continue
    # Split by | and strip
    cells = [c.strip() for c in line.split("|")]
    # Remove empty first and last from split
    cells = [c for c in cells if c or (cells.index(c) > 0 and cells.index(c) < len(cells)-1)]
    # Actually, just keep all non-edge empties
    cells_raw = line.split("|")
    cells = [c.strip() for c in cells_raw[1:-1]]  # skip first and last empty from leading/trailing |
    
    if len(cells) >= 18:
        rows.append(cells[:18])

print(f"Parsed {len(rows)} rows from table")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = f"W20文献调研"

# Headers
headers = [
    "序号", "调研日期", "标题", "作者", "期刊", "发表日期",
    "DOI", "发表状态", "研究方向", "关键词编号", "核心发现",
    "优先级", "JCR分区", "影响因子", "检索来源", "备注"
]

# Column mapping from source to output
# Source columns: #=0, 标题=1, 作者=2, 期刊=3, 发表日期=4, DOI=5, 发表状态=6, 摘要=7, 论文关键词=8, 核心发现=9, 检索来源=10, 关键词编号=11, 检索组=12, 研究方向=13, 优先级=14, JCR分区=15, 影响因子=16, 备注=17

# Style definitions
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="微软雅黑", size=10)
cell_alignment = Alignment(vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data rows
for row_idx, row_data in enumerate(rows, 2):
    seq = row_data[0]
    title = row_data[1]
    authors = row_data[2]
    journal = row_data[3]
    pub_date = row_data[4]
    doi = row_data[5]
    pub_status = row_data[6]
    # abstract = row_data[7]  # Not in output columns
    # keywords = row_data[8]  # Not in output columns
    core_finding = row_data[9]
    source = row_data[10]
    kw_num = row_data[11]
    # group = row_data[12]  # Not in output columns
    direction = row_data[13]
    priority = row_data[14]
    jcr = row_data[15]
    impact_factor = row_data[16]
    note = row_data[17] if len(row_data) > 17 else ""
    
    # Output column order: 序号, 调研日期, 标题, 作者, 期刊, 发表日期, DOI, 发表状态, 研究方向, 关键词编号, 核心发现, 优先级, JCR分区, 影响因子, 检索来源, 备注
    output_values = [
        int(seq) if seq.isdigit() else seq,
        SURVEY_DATE,
        title,
        authors,
        journal,
        pub_date,
        doi,
        pub_status,
        direction,
        kw_num,
        core_finding,
        priority,
        jcr if jcr and jcr != "N/A" else "",
        impact_factor if impact_factor and impact_factor != "N/A" else "",
        source,
        note if note and note != "N/A" else ""
    ]
    
    for col_idx, value in enumerate(output_values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        
        # Center alignment for specific columns
        if col_idx in [1, 2, 6, 8, 10, 12, 13, 14, 15]:
            cell.alignment = center_alignment
        else:
            cell.alignment = cell_alignment

# Auto-adjust column widths
col_widths = {
    1: 6,    # 序号
    2: 12,   # 调研日期
    3: 45,   # 标题
    4: 25,   # 作者
    5: 20,   # 期刊
    6: 12,   # 发表日期
    7: 30,   # DOI
    8: 14,   # 发表状态
    9: 25,   # 研究方向
    10: 12,  # 关键词编号
    11: 50,  # 核心发现
    12: 8,   # 优先级
    13: 8,   # JCR分区
    14: 10,  # 影响因子
    15: 22,  # 检索来源
    16: 30,  # 备注
}

for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Add auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Ensure output directory exists
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

# Save
wb.save(OUTPUT_FILE)
print(f"Excel saved to: {OUTPUT_FILE}")
print(f"Total papers: {len(rows)}")
